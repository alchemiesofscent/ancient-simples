#!/usr/bin/env python3
from __future__ import annotations

import csv
import re
import sys
import unicodedata
from collections import defaultdict, OrderedDict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


GREEK_TOKEN_RE = re.compile(r"[\u0370-\u03FF\u1F00-\u1FFF]+", re.UNICODE)


@dataclass(frozen=True)
class LemmaRecord:
    lemma_id: str
    headword_gr: str
    headword_normalized: str
    headword_en: str
    parent_lemma: str
    relationship: str
    category: str
    notes: str


def normalize_greek_for_match(text: str) -> str:
    text = text.strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(
        ch for ch in text if not unicodedata.combining(ch) or ch == "\u0345"
    )
    return unicodedata.normalize("NFC", text)


def is_nullish_cell(value: str) -> bool:
    v = value.strip().lower()
    return v in {"", "null", "(null)", "nan"}


def split_lemma_items(value: str) -> list[tuple[str, str]]:
    """
    Split a single workbook lemma cell into ordered lemma items.

    Returns list of (item, notes) where notes is either "" or the required
    "split from workbook list: <original cell>" text.
    """
    original = str(value).strip()
    if is_nullish_cell(original):
        return []

    # Base split: commas + semicolons always.
    base_parts = [p.strip() for p in re.split(r"[;,]", original) if p.strip()]

    original_norm = normalize_greek_for_match(original)
    kai_count = original_norm.count(" και ")
    te_kai_count = original_norm.count(" τε και ")
    list_like = ("," in original or ";" in original) or (kai_count + te_kai_count >= 3)

    items: list[str] = []
    if list_like:
        # Further split on conjunctions only when the cell clearly encodes a list.
        conj_re = re.compile(r"\s+(?:τε\s+)?κα(?:ὶ|ι)\s+", re.UNICODE)
        for part in base_parts:
            subparts = [s.strip() for s in conj_re.split(part) if s.strip()]
            items.extend(subparts if subparts else [part])
    else:
        items = base_parts

    items = [i for i in items if i.strip()]
    if len(items) <= 1:
        return [(items[0], "")] if items else []

    note = f"split from workbook list: {original}"
    return [(item, note) for item in items]


def first_greek_token(text: str) -> str | None:
    m = GREEK_TOKEN_RE.search(text)
    return m.group(0) if m else None


def looks_like_abstract(norm: str) -> bool:
    abstract_stems = [
        "διαφορ",
        "ενεργει",
        "ιδιοτ",
        "πασχ",
        "θεωρι",
        "λογος",
        "ονομα",
    ]
    return any(stem in norm for stem in abstract_stems)


def category_for_lemma(norm: str) -> str:
    # Conservative: default to broad "substance" unless we are confident.
    if looks_like_abstract(norm):
        return "abstract"

    animal_stems = [
        "αιμα",
        "γαλα",
        "χολ",
        "κοπρ",
        "ουρ",
        "ηπαρ",
        "οστο",
        "κερα",
        "δερμ",
        "ονυ",
        "ωον",
        "πιμελ",
        "μυελ",
    ]
    if any(norm.startswith(stem) for stem in animal_stems):
        return "animal"

    mineral_stems = [
        "λιθ",
        "γη",
        "γυψ",
        "ασβεστ",
        "χαλκ",
        "σιδηρ",
        "αργυρ",
        "χρυσ",
        "μολυβ",
        "θει",
        "μεταλλ",
    ]
    if any(stem in norm for stem in mineral_stems):
        return "mineral"

    product_stems = [
        "τεφρ",
        "σποδ",
        "σποδι",
    ]
    if any(norm.startswith(stem) for stem in product_stems):
        return "product"

    return "substance"


def has_non_greek_letters(headword_gr: str) -> bool:
    # Allow spaces/punctuation, but flag Latin letters/digits.
    for ch in headword_gr:
        if "0" <= ch <= "9":
            return True
        if "A" <= ch <= "Z" or "a" <= ch <= "z":
            return True
    return False


def main() -> int:
    repo_root = Path(__file__).resolve().parents[1]
    workbench = repo_root / "data-workbench"
    xlsx_path = workbench / "Simples.xlsx"

    lemmata_csv_path = workbench / "lemmata.csv"
    lemmata_review_csv_path = workbench / "lemmata_review.csv"
    lemmata_qc_md_path = workbench / "lemmata_qc.md"
    column_o_audit_md_path = workbench / "columnO_category_audit.md"

    if not xlsx_path.exists():
        print(f"ERROR: missing input workbook: {xlsx_path}", file=sys.stderr)
        return 2

    xl = pd.ExcelFile(xlsx_path, engine="openpyxl")

    # Column O (Excel) is authoritative category for lemmata.
    # Map raw values to the controlled category set used in lemmata.csv.
    category_map = {
        "plant": "plant",
        "Plant": "plant",
        "Animal": "animal",
        "Mineral": "mineral",
    }
    column_o_blank_fallbacks = 0
    column_o_unmapped_counts: dict[str, int] = defaultdict(int)

    review_whitelist_norm = {"εψησις", "εψησεως", "βρεξαντα"}

    # Stable incremental IDs in worksheet order: first occurrence wins insertion order.
    headword_first_context: dict[str, tuple[str, int, str]] = {}
    headword_first_note: dict[str, str] = {}
    headword_first_category: dict[str, str] = {}
    ordered_headwords: "OrderedDict[str, None]" = OrderedDict()
    review_rows: list[dict[str, str]] = []

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    # 1) Column O audit (header + distinct values).
    column_o_counts: dict[str, int] = defaultdict(int)
    audit_lines: list[str] = []
    audit_lines.append("# Column O (Category) audit")
    audit_lines.append("")
    audit_lines.append(f"- Workbook: `{xlsx_path.name}`")
    audit_lines.append(f"- Sheets: {', '.join(f'`{s}`' for s in wb.sheetnames)}")
    audit_lines.append("")
    audit_lines.append("## Sheet headers")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        audit_lines.append(
            f"- `{sheet_name}`: lemma column `M1` = `{ws['M1'].value}`, category column `O1` = `{ws['O1'].value}`"
        )
    audit_lines.append("")

    # 2) Collect lemmata in worksheet order using column M (Lemma) and column O (Category).
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        row_index_1 = 1
        for row_index_1 in range(2, ws.max_row + 1):
            lemma_cell = ws[f"M{row_index_1}"].value
            category_raw = ws[f"O{row_index_1}"].value

            if lemma_cell is None:
                continue
            raw = str(lemma_cell)

            category_raw_s = "" if category_raw is None else str(category_raw).strip()
            if category_raw_s:
                column_o_counts[category_raw_s] += 1

            if not category_raw_s:
                category = "substance"
                column_o_blank_fallbacks += 1
            else:
                category = category_map.get(category_raw_s)
                if not category:
                    for item, _notes in split_lemma_items(raw):
                        if is_nullish_cell(item):
                            continue
                        review_rows.append(
                            {
                                "headword_gr": item,
                                "context": f"{sheet_name} row {row_index_1}: {raw}",
                                "reason": f"unknown category value in column O: {category_raw_s}",
                                "suggested_category_or_parent": "",
                            }
                        )
                    column_o_unmapped_counts[category_raw_s] += 1
                    continue

            for item, notes in split_lemma_items(raw):
                if is_nullish_cell(item):
                    continue

                item_norm = normalize_greek_for_match(item)
                if item_norm in review_whitelist_norm:
                    review_rows.append(
                        {
                            "headword_gr": item,
                            "context": f"{sheet_name} row {row_index_1}: {raw}",
                            "reason": "verbal/process form; belongs in preparations/process layer",
                            "suggested_category_or_parent": "preparations/process",
                        }
                    )
                    continue

                # Persist first-seen category deterministically; if the first-seen category
                # was a fallback from blank column O, allow overwrite by a later non-blank.
                if item not in ordered_headwords:
                    ordered_headwords[item] = None
                    headword_first_context[item] = (sheet_name, row_index_1, raw)
                    note_out = notes
                    if notes:
                        note_out = (note_out + "; " if note_out else "") + "category inherited from bundle source"
                    headword_first_note[item] = note_out
                    headword_first_category[item] = category
                else:
                    if notes and not headword_first_note.get(item):
                        note_out = notes
                        note_out = (note_out + "; " if note_out else "") + "category inherited from bundle source"
                        headword_first_note[item] = note_out
                    existing_cat = headword_first_category.get(item)
                    if existing_cat == "substance" and category_raw_s and category != "substance":
                        headword_first_category[item] = category

    # Write audit file.
    audit_lines.append("## Distinct column O values (trimmed)")
    for v, c in sorted(column_o_counts.items(), key=lambda kv: (-kv[1], kv[0]))[:30]:
        audit_lines.append(f"- `{v}`: {c}")
    audit_lines.append("")
    column_o_audit_md_path.write_text("\n".join(audit_lines), encoding="utf-8", newline="\n")

    headwords = list(ordered_headwords.keys())

    lemma_ids: dict[str, str] = {}
    records: list[LemmaRecord] = []
    # review_rows already populated above.

    # Index single-token lemmata by normalized first token for conservative parent assignment.
    single_token_norm_to_ids: dict[str, list[str]] = defaultdict(list)

    for i, headword_gr in enumerate(headwords, start=1):
        lemma_id = f"L{i:03d}"
        lemma_ids[headword_gr] = lemma_id

        headword_norm = normalize_greek_for_match(headword_gr)
        category = headword_first_category.get(headword_gr, "substance")

        notes = headword_first_note.get(headword_gr, "") or ""

        records.append(
            LemmaRecord(
                lemma_id=lemma_id,
                headword_gr=headword_gr,
                headword_normalized=headword_norm,
                headword_en="",
                parent_lemma="",
                relationship="",
                category=category,
                notes=notes,
            )
        )

        tokens = GREEK_TOKEN_RE.findall(headword_gr)
        if len(tokens) == 1:
            single_token_norm_to_ids[normalize_greek_for_match(tokens[0])].append(lemma_id)

        # Review flags (conservative).
        sheet, row_index_1, raw = headword_first_context.get(headword_gr, ("", 0, ""))
        if has_non_greek_letters(headword_gr):
            review_rows.append(
                {
                    "headword_gr": headword_gr,
                    "context": f"{sheet} row {row_index_1}: {raw}",
                    "reason": "Contains non-Greek letters/digits.",
                    "suggested_category_or_parent": f"category={category}",
                }
            )
        if "(" in headword_gr or ")" in headword_gr:
            review_rows.append(
                {
                    "headword_gr": headword_gr,
                    "context": f"{sheet} row {row_index_1}: {raw}",
                    "reason": "Contains parentheses; may encode variant/aside.",
                    "suggested_category_or_parent": "",
                }
            )

    # Second pass: parent_lemma + relationship where explicit modifier patterns appear.
    prep_modifier_norms = {
        "κεκαυμενος",
        "κεκαυμενη",
        "κεκαυμενον",
        "κεκαυμενοι",
        "κεκαυμεναι",
        "κεκαυμενα",
        "κεκαυμενων",
        "κεκαυμενους",
        "κεκαυμενης",
        "κεκαυμενου",
    }
    explicit_variant_norms = {
        "μεγα",
        "μικρον",
        "ημερος",
        "αγριος",
        "αγρια",
        "διττη",
        "πασα",
    }
    derivation_markers_norm = {"εκ", "απο", "απ"}

    updated_records: list[LemmaRecord] = []
    for rec in records:
        headword_gr = rec.headword_gr
        tokens = GREEK_TOKEN_RE.findall(headword_gr)
        parent_lemma = ""
        relationship = ""

        if len(tokens) >= 2:
            base_token = tokens[0]
            base_norm = normalize_greek_for_match(base_token)
            parent_candidates = single_token_norm_to_ids.get(base_norm, [])
            if len(parent_candidates) == 1:
                # Only assign parent when the remainder explicitly signals derivation/variant/modifier.
                remainder_norms = [normalize_greek_for_match(t) for t in tokens[1:]]
                if any(n in prep_modifier_norms for n in remainder_norms):
                    parent_lemma = parent_candidates[0]
                    relationship = "subtype"
                elif any(n in explicit_variant_norms for n in remainder_norms):
                    parent_lemma = parent_candidates[0]
                    relationship = "subtype"
                elif remainder_norms and remainder_norms[0] in derivation_markers_norm:
                    parent_lemma = parent_candidates[0]
                    relationship = "subtype"
            elif len(parent_candidates) > 1:
                sheet, row_index_1, raw = headword_first_context.get(headword_gr, ("", 0, ""))
                review_rows.append(
                    {
                        "headword_gr": headword_gr,
                        "context": f"{sheet} row {row_index_1}: {raw}",
                        "reason": "Ambiguous parent candidate (multiple matching base lemmata).",
                        "suggested_category_or_parent": f"parent_base_norm={base_norm}",
                    }
                )

        updated_records.append(
            LemmaRecord(
                lemma_id=rec.lemma_id,
                headword_gr=rec.headword_gr,
                headword_normalized=rec.headword_normalized,
                headword_en=rec.headword_en,
                parent_lemma=parent_lemma,
                relationship=relationship,
                category=rec.category,
                notes=rec.notes,
            )
        )

    # A) Parent/subtype enrichment for oil family (ελαιον).
    oil_parent_id = ""
    for r in updated_records:
        if r.headword_normalized == "ελαιον":
            oil_parent_id = r.lemma_id
            break

    parent_assigned_A = 0
    if oil_parent_id:
        tmp: list[LemmaRecord] = []
        for r in updated_records:
            if r.headword_normalized != "ελαιον" and (
                r.headword_normalized.endswith(" ελαιον")
                or r.headword_normalized.startswith("ελαιον ")
            ):
                if not r.parent_lemma:
                    parent_assigned_A += 1
                    tmp.append(
                        LemmaRecord(
                            lemma_id=r.lemma_id,
                            headword_gr=r.headword_gr,
                            headword_normalized=r.headword_normalized,
                            headword_en=r.headword_en,
                            parent_lemma=oil_parent_id,
                            relationship="subtype",
                            category=r.category,
                            notes=r.notes,
                        )
                    )
                    continue
            tmp.append(r)
        updated_records = tmp

    # D) Conservative parent-prefix rule: multiword lemma "X ..." inherits parent X when available.
    single_token_norm_to_id: dict[str, str] = {}
    for r in updated_records:
        if " " not in r.headword_normalized.strip():
            single_token_norm_to_id.setdefault(r.headword_normalized, r.lemma_id)

    parent_assigned_D = 0
    tmp2: list[LemmaRecord] = []
    for r in updated_records:
        if r.parent_lemma or " " not in r.headword_normalized:
            tmp2.append(r)
            continue
        base = r.headword_normalized.split(" ", 1)[0]
        parent = single_token_norm_to_id.get(base)
        if parent:
            parent_assigned_D += 1
            tmp2.append(
                LemmaRecord(
                    lemma_id=r.lemma_id,
                    headword_gr=r.headword_gr,
                    headword_normalized=r.headword_normalized,
                    headword_en=r.headword_en,
                    parent_lemma=parent,
                    relationship="subtype",
                    category=r.category,
                    notes=r.notes,
                )
            )
        else:
            tmp2.append(r)
    updated_records = tmp2

    lemmata_df = pd.DataFrame([r.__dict__ for r in updated_records])[
        [
            "lemma_id",
            "headword_gr",
            "headword_normalized",
            "headword_en",
            "parent_lemma",
            "relationship",
            "category",
            "notes",
        ]
    ]
    lemmata_df.to_csv(
        lemmata_csv_path,
        index=False,
        encoding="utf-8",
        quoting=csv.QUOTE_MINIMAL,
        lineterminator="\n",
    )

    review_df = pd.DataFrame(
        review_rows, columns=["headword_gr", "context", "reason", "suggested_category_or_parent"]
    )
    review_df.to_csv(
        lemmata_review_csv_path,
        index=False,
        encoding="utf-8",
        quoting=csv.QUOTE_MINIMAL,
        lineterminator="\n",
    )

    # QC report
    utc_now = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    norm_collisions: dict[str, list[str]] = defaultdict(list)
    iota_subscript_count = 0
    non_greek_count = 0
    multiword_count = 0
    for r in updated_records:
        norm_collisions[r.headword_normalized].append(r.headword_gr)
        if "\u0345" in unicodedata.normalize("NFD", r.headword_normalized):
            iota_subscript_count += 1
        if has_non_greek_letters(r.headword_gr):
            non_greek_count += 1
        if len(GREEK_TOKEN_RE.findall(r.headword_gr)) >= 2:
            multiword_count += 1

    collisions = {k: v for k, v in norm_collisions.items() if len(v) >= 2}
    category_counts: dict[str, int] = defaultdict(int)
    for r in updated_records:
        category_counts[r.category] += 1

    qc_lines: list[str] = []
    qc_lines.append("# lemmata.csv QC report")
    qc_lines.append("")
    qc_lines.append(f"- Generated: `{utc_now}`")
    qc_lines.append(f"- Workbook: `{xlsx_path.name}`")
    qc_lines.append(f"- Total lemmata: **{len(lemmata_df)}**")
    qc_lines.append(f"- Sent to review: **{len(review_df)}**")
    qc_lines.append("")
    qc_lines.append("## Normalization edge cases")
    qc_lines.append(f"- Headwords containing iota subscripts (preserved): **{iota_subscript_count}**")
    qc_lines.append(f"- Headwords containing non-Greek letters/digits: **{non_greek_count}**")
    qc_lines.append(f"- Multiword headwords (>=2 Greek tokens): **{multiword_count}**")
    qc_lines.append(f"- Normalized collisions (distinct headwords sharing same normalized form): **{len(collisions)}**")
    if collisions:
        shown = 0
        for norm, variants in sorted(collisions.items(), key=lambda kv: (-len(kv[1]), kv[0])):
            qc_lines.append(f"  - `{norm}`: {', '.join(f'`{v}`' for v in variants[:6])}{'…' if len(variants) > 6 else ''}")
            shown += 1
            if shown >= 20:
                qc_lines.append("  - _(more omitted)_")
                break
    qc_lines.append("")
    qc_lines.append("## Category (from column O)")
    for cat, count in sorted(category_counts.items(), key=lambda kv: (-kv[1], kv[0])):
        qc_lines.append(f"- `{cat}`: **{count}**")
    qc_lines.append(f"- Column O blank fallbacks to `substance`: **{column_o_blank_fallbacks}**")
    unmapped_total = sum(column_o_unmapped_counts.values())
    qc_lines.append(f"- Sent to review for unmapped column O values: **{unmapped_total}**")
    if column_o_unmapped_counts:
        for raw, count in sorted(column_o_unmapped_counts.items(), key=lambda kv: (-kv[1], kv[0])):
            qc_lines.append(f"  - `{raw}`: {count}")
    qc_lines.append("")
    qc_lines.append("## Parent assignment")
    qc_lines.append(f"- Oil family (A) parent assignments: **{parent_assigned_A}**")
    qc_lines.append(f"- Conservative prefix (D) parent assignments: **{parent_assigned_D}**")
    qc_lines.append("")

    lemmata_qc_md_path.write_text("\n".join(qc_lines), encoding="utf-8", newline="\n")

    print(f"Wrote {lemmata_csv_path} ({len(lemmata_df)} rows)")
    print(f"Wrote {lemmata_review_csv_path} ({len(review_df)} rows)")
    print(f"Parent assignments: A={parent_assigned_A}, D={parent_assigned_D}")
    print(f"Wrote {lemmata_qc_md_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
